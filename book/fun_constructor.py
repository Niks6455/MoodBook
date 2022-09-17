import openpyxl

name_book = None
name_avtor = None
izdatel = None
god_vipuska = None
janr = None
vozrast = None
opisanie = None
reccenzia = None
avtor_reccenzia = None
silka = None
count = None
namber = None

s = None
s1 = None
sheet = None
wb = None
ranCount = None
xl = "D:/WebsiteProgect/modbook/book/inform.xlsx"


#------------------------
#--Считаем данные из xl--
#------------------------
def xlread():
    global name_book
    global name_avtor
    global izdatel
    global god_vipuska
    global janr 
    global vozrast 
    global opisanie 
    global reccenzia
    global avtor_reccenzia
    global silka
    global count
    global sheet
    global wb
    global namber
    wb = openpyxl.reader.excel.load_workbook(filename=xl)
    #print(wb.sheetnames)
    wb.active = 0 #установили активным лист по индексу 0 из xl
    sheet = wb.active
    #print('\n')
    #for i in range(1,11):
    #    print(sheet["A"+str(i)].value + '\t' + str(sheet["B"+str(i)].value))
    #print('\n')
    name_book = str(sheet["B1"].value)
    name_avtor = str(sheet["B2"].value)
    izdatel = str(sheet["B3"].value)
    god_vipuska = str(sheet["B4"].value)
    janr = str(sheet["B5"].value)
    vozrast = str(sheet["B6"].value)
    opisanie = str(sheet["B7"].value)
    reccenzia = str(sheet["B8"].value)
    avtor_reccenzia = str(sheet["B9"].value)
    silka = str(sheet["B10"].value)
    count = str(sheet["D1"].value)
    namber = str(sheet["C6"].value)

def count_sum():
    global sheet
    global wb
    sheet['D1'].value = str(int(sheet["D1"].value) + 1)
    wb.save(xl)


# def ran():
#     global ranCount
#     wb = openpyxl.reader.excel.load_workbook(filename=xl)
#     wb.active = 0 #установили активным лист по индексу 0 из xl
#     sheet = wb.active
#     ranCount = str(sheet['E9'].value)
#     wb.save(xl)
#     print("E9 = " + str(ranCount))

# def endran():
#     global sheet
#     global wb
#     sheet['E9'].value = str(0)
#     wb.save(xl)
#     print("E9 = 0")

#------------------------
#---Считаем book.html---
#------------------------
def disk(lin,newstroca): # функция перезаписи
        #print(lin)
        #print(s[lin])
        s[lin] = newstroca
        #print(s[lin])
        #print('\n')

def book_read():
    global s
    fileBook = open("D:/WebsiteProgect/modbook/book/book.html", "r" , encoding='utf-8' )
    s = fileBook.readlines()
    fileBook.close()

    lin = s.index("              <div class=\"book__name\">Унесённые ветром</div>\n")
    newstroca = "              <div class=\"book__name\">" + name_book + "</div>\n"
    disk(lin,newstroca)

    lin = s.index("                  src=\"../assets/image/обложки/унесенные ветром.jpg\"\n")
    newstroca = "                  src=\"" + silka + "\"\n"
    disk(lin,newstroca)

    lin = s.index("                        <span>Автор:</span>Маргарет Митчел\n")
    newstroca = "                        <span>Автор:</span>"+name_avtor+"\n"
    disk(lin,newstroca)

    lin = s.index("                        <span>Издательство:</span>АСТ\n")
    newstroca = "                        <span>Издательство:</span>" + izdatel+"\n"
    disk(lin,newstroca)

    lin = s.index("                      <div class=\"book__age\"><span>Год выпуска:</span>1936</div>\n")
    newstroca = "                      <div class=\"book__age\"><span>Год выпуска:</span>" + god_vipuska + "</div>\n"
    disk(lin,newstroca)

    lin = s.index("                      <div class=\"book__ganre\"><span>Жанр:</span>Роман</div>\n")
    newstroca = "                      <div class=\"book__ganre\"><span>Жанр:</span>" + janr + "</div>\n"
    disk(lin,newstroca)

    lin = s.index("                      <div class=\"book__ograge\">16+</div>\n")
    newstroca = "                      <div class=\"book__ograge\">" + vozrast + "</div>\n"
    disk(lin,newstroca)

    lin = s.index("                <div class=\"book__description\">Описание</div>\n")
    newstroca = "                <div class=\"book__description\">" + opisanie + "</div>\n"
    disk(lin,newstroca)

    lin = s.index("                <span>Автор:</span>Юлия.К\n")
    newstroca = "                <span>Автор:</span>" + avtor_reccenzia + "\n"
    disk(lin,newstroca)

    lin = s.index("                <div class=\"rec__text\">Рецензия</div>\n")
    newstroca = "                <div class=\"rec__text\">" + reccenzia + "</div>\n"
    disk(lin,newstroca)


#-----------------------------------
#---Создадим новый файл и запишем---
#-----------------------------------

def new_book_write():
    href = "D:/WebsiteProgect/modbook/book/book" + count + ".html"
    file2 = open(href, "w" , encoding='utf-8' )
    file2.write("")
    file2.close()

    file2 = open(href, "a" , encoding='utf-8' )
    for i in range(0,len(s)):
        file2.write(s[i])
    print("Файл записан")
    file2.close()




#-----------------------------------
#---Подключим книгу vse_book.html---
#-----------------------------------
def vse_book_write():
    global s1
    global janr_class
    fileVseBook = open("D:/WebsiteProgect/modbook/index/vse__book.html", "r" , encoding='utf-8' )
    s1 = fileVseBook.readlines()
    fileVseBook.close()

    if namber == str(1):
        janr_class = "novels"
    if namber == str(2):
        janr_class = "fantasy"
    if namber == str(3):
        janr_class = "artistic"
    if namber == str(4):
        janr_class = "science"
    if namber == str(5):
        janr_class = "psychology"
    if namber == str(6):
        janr_class = "classic"
    if namber == str(7):
        janr_class = "motivation"


    st = "        <div class=\"chto_block " + janr_class+"\">\n          <div class=\"chto_obloshka\">\n            <img src=\""+silka+"\">\n          </div>\n          <div class=\"chto_block_name\">"+name_book+"</div>\n          <div class=\"chto_block_avtor\">"+ name_avtor+"</div>\n          <div class=\"chto_god\">Год издания:"+god_vipuska+"</br> Жанр:"+janr +"</div>\n          <div class=\"chto_star\">\n            <img src=\"../assets/image/logo/star.png\"> 4.4\n          </div>\n          <a href=\"../book/book"+count+".html\" class=\"chto_button\">\n            <div class=\"chto_podrob\"> Подробнее</div>\n          </a>\n        </div>\n\n\n"

    index = s1.index('  <div class="pop">\n') - 2
    s1[index] = st

    fileVseBook_write = open("D:/WebsiteProgect/modbook/index/vse__book.html", "w" , encoding='utf-8' )
    for i in range(0,len(s1)):
        fileVseBook_write.write(s1[i])
    print("Файл все книги записан")
    fileVseBook_write.close()


#-----------------------------------
#---Подключим книгу podb.html---
#-----------------------------------
def podb_write():
    global s1
    fileVseBook = open("D:/WebsiteProgect/modbook/index/podb6.html", "r" , encoding='utf-8' )
    s1 = fileVseBook.readlines()
    fileVseBook.close()

    st = '\n              <div class="books novels">\n                <a href="../book/book'+count+'.html"><img class="bg" src="'+silka+'"/></a>\n                <div class="books__text">\n                  <div class="books__name"><a href="#">'+name_book+'</a></div>\n                  <div class="books__aphtor"><a href="#">'+name_avtor+'</a></div>\n                </div>\n              </div>\n\n\n'

    index = s1.index('      <!-- footer -->\n') - 5
    s1.insert(index,st)

    fileVseBook_write = open("D:/WebsiteProgect/modbook/index/podb6.html", "w" , encoding='utf-8' )
    for i in range(0,len(s1)):
        fileVseBook_write.write(s1[i])
    print("Файл podb4.html записан")
    fileVseBook_write.close()
